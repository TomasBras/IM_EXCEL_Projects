from typing import Any, Text, Dict, List
from rasa_sdk import Action, Tracker
from rasa_sdk.executor import CollectingDispatcher
from rasa_sdk.events import SlotSet, UserUtteranceReverted
import pandas as pd
import openpyxl
import json
import os
import matplotlib.pyplot as plt


def write_log(text: str):
    with open("log.txt", "a", encoding="utf-8") as log:
        log.write(text + "\n")


EXCEL_PATH = r"C:\Users\trmbr\OneDrive\Desktop\IM\IM_EXCEL_NODEPENDENCIES\rasaDemo\dados_turma.xlsx"

class ActionCalcularMedia(Action):
    def name(self) -> Text:
        return "action_calcular_media"

    def run(self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        aluno = tracker.get_slot("aluno")
        write_log("Ação: Calcular Média - Início")

        try:
            df = pd.read_excel(EXCEL_PATH)
            if aluno:
                if aluno in df["Aluno"].values:
                    notas = df.loc[df["Aluno"] == aluno, ["Teste1", "Teste2", "Teste3"]].values.flatten()
                    media = round(sum(notas) / len(notas), 2)
                    df.loc[df["Aluno"] == aluno, "Média"] = media
                    dispatcher.utter_message(text=f"A média de {aluno} foi calculada: {media}.")
                else:
                    dispatcher.utter_message(text=f"Não encontrei o aluno {aluno} na folha.")
            else:
                df["Média"] = df[["Teste1", "Teste2", "Teste3"]].mean(axis=1)
                dispatcher.utter_message(text="A média da turma foi calculada com sucesso.")

            df.to_excel(EXCEL_PATH, index=False)
            write_log("Ação: Calcular Média - Sucesso")
        except Exception as e:
            dispatcher.utter_message(text="Ocorreu um erro ao calcular a média.")
            write_log(f"Erro em Calcular Média: {str(e)}")

        return []


class ActionDestacarAprovadosReprovados(Action):
    def name(self) -> Text:
        return "action_destacar_aprovados_reprovados"

    def run(self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        write_log("Ação: Destacar Aprovados/Reprovados - Início")

        try:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb.active
            for cell in ws["D"][1:]:  # Coluna D = Média (ajustar se necessário)
                try:
                    if float(cell.value) >= 10:
                        cell.fill = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    else:
                        cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                except:
                    pass
            wb.save(EXCEL_PATH)
            dispatcher.utter_message(text="Os alunos aprovados e reprovados foram destacados com cores.")
            write_log("Ação: Destacar Aprovados/Reprovados - Sucesso")
        except Exception as e:
            dispatcher.utter_message(text="Erro ao aplicar formatação condicional.")
            write_log(f"Erro em Destacar Aprovados/Reprovados: {str(e)}")

        return []


class ActionIdentificarMelhoria(Action):
    def name(self) -> Text:
        return "action_identificar_melhoria"

    def run(self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        write_log("Ação: Identificar Melhoria - Início")

        try:
            df = pd.read_excel(EXCEL_PATH)
            df["Melhoria"] = ((df["Teste3"] - df["Teste1"]) / df["Teste1"]) * 100
            melhoraram = df[df["Melhoria"] > 20]
            dispatcher.utter_message(text=f"{len(melhoraram)} alunos tiveram melhoria superior a 20%.")
            df.to_excel(EXCEL_PATH, index=False)
            write_log("Ação: Identificar Melhoria - Sucesso")
        except Exception as e:
            dispatcher.utter_message(text="Erro ao calcular melhorias.")
            write_log(f"Erro em Identificar Melhoria: {str(e)}")

        return []


class ActionInserirColunas(Action):
    def name(self) -> Text:
        return "action_inserir_colunas"

    def run(self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        write_log("Ação: Inserir Coluna Situação - Início")

        try:
            df = pd.read_excel(EXCEL_PATH)
            if "Situação" not in df.columns:
                df["Situação"] = df["Média"].apply(lambda x: "Aprovado" if x >= 10 else "Reprovado")
                dispatcher.utter_message(text="Coluna 'Situação' criada com sucesso.")
            else:
                dispatcher.utter_message(text="A coluna 'Situação' já existe.")
            df.to_excel(EXCEL_PATH, index=False)
            write_log("Ação: Inserir Coluna Situação - Sucesso")
        except Exception as e:
            dispatcher.utter_message(text="Erro ao inserir a coluna Situação.")
            write_log(f"Erro em Inserir Colunas: {str(e)}")

        return []


class ActionGerarGraficos(Action):
    def name(self) -> Text:
        return "action_gerar_graficos"

    def run(self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        write_log("Ação: Gerar Gráficos - Início")

        try:
            df = pd.read_excel(EXCEL_PATH)
            plt.figure()
            df.plot(x="Aluno", y=["Teste1", "Teste2", "Teste3"], kind="line", marker="o")
            plt.title("Evolução de Desempenho dos Alunos")
            plt.ylabel("Notas")
            plt.savefig("grafico_evolucao.png")
            dispatcher.utter_message(text="Gráfico de evolução criado com sucesso.")
            write_log("Ação: Gerar Gráficos - Sucesso")
        except Exception as e:
            dispatcher.utter_message(text="Erro ao gerar o gráfico.")
            write_log(f"Erro em Gerar Gráficos: {str(e)}")

        return []


class ActionAtualizarNotas(Action):
    def name(self) -> Text:
        return "action_atualizar_notas"

    def run(self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        aluno = tracker.get_slot("aluno")
        valores = tracker.get_slot("valores")
        write_log("Ação: Atualizar Notas - Início")

        try:
            valores_lista = [float(x.strip()) for x in valores.strip("[]").split(",")]
            df = pd.read_excel(EXCEL_PATH)
            if aluno in df["Aluno"].values:
                df.loc[df["Aluno"] == aluno, ["Teste1", "Teste2", "Teste3"]] = valores_lista
                df.to_excel(EXCEL_PATH, index=False)
                dispatcher.utter_message(text=f"Notas de {aluno} atualizadas para {valores_lista}.")
                write_log("Ação: Atualizar Notas - Sucesso")
            else:
                dispatcher.utter_message(text=f"Aluno {aluno} não encontrado.")
        except Exception as e:
            dispatcher.utter_message(text="Erro ao atualizar as notas.")
            write_log(f"Erro em Atualizar Notas: {str(e)}")

        return []


class ActionGuardarFicheiro(Action):
    def name(self) -> Text:
        return "action_guardar_ficheiro"

    def run(self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        nome_ficheiro = tracker.get_slot("nome_ficheiro")
        write_log("Ação: Guardar Ficheiro - Início")

        try:
            if os.path.exists(EXCEL_PATH):
                novo_nome = nome_ficheiro or "Relatório_Final.xlsx"
                os.rename(EXCEL_PATH, novo_nome)
                dispatcher.utter_message(text=f"Ficheiro guardado como {novo_nome}.")
                write_log("Ação: Guardar Ficheiro - Sucesso")
            else:
                dispatcher.utter_message(text="Não encontrei o ficheiro de dados.")
        except Exception as e:
            dispatcher.utter_message(text="Erro ao guardar o ficheiro.")
            write_log(f"Erro em Guardar Ficheiro: {str(e)}")

        return []


class ActionOperacoesMatematicas(Action):
    def name(self) -> Text:
        return "action_operacoes_matematicas"

    def run(self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        write_log("Ação: Operações Matemáticas - Início")

        try:
            df = pd.read_excel(EXCEL_PATH)
            aprovados = len(df[df["Média"] >= 10])
            reprovados = len(df[df["Média"] < 10])
            percentagem = round((aprovados / len(df)) * 100, 2)
            dispatcher.utter_message(
                text=f"{aprovados} alunos aprovados ({percentagem}%) e {reprovados} reprovados."
            )
            write_log("Ação: Operações Matemáticas - Sucesso")
        except Exception as e:
            dispatcher.utter_message(text="Erro ao realizar as operações.")
            write_log(f"Erro em Operações Matemáticas: {str(e)}")

        return []


class ActionAtualizarValores(Action):
    def name(self) -> Text:
        return "action_atualizar_valores"

    def run(self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        aluno = tracker.get_slot("aluno")
        pergunta = tracker.get_slot("pergunta")
        valor = tracker.get_slot("valor_novo")
        write_log("Ação: Atualizar Valores - Início")

        try:
            df = pd.read_excel(EXCEL_PATH)
            col = f"Pergunta {pergunta}"
            if col in df.columns:
                if aluno and aluno in df["Aluno"].values:
                    df.loc[df["Aluno"] == aluno, col] = float(valor)
                    dispatcher.utter_message(text=f"Valor da {col} atualizado para {aluno}.")
                else:
                    df[col] = float(valor)
                    dispatcher.utter_message(text=f"Valor da {col} atualizado para toda a turma.")
                df.to_excel(EXCEL_PATH, index=False)
                write_log("Ação: Atualizar Valores - Sucesso")
            else:
                dispatcher.utter_message(text=f"A coluna {col} não existe no ficheiro.")
        except Exception as e:
            dispatcher.utter_message(text="Erro ao atualizar valores.")
            write_log(f"Erro em Atualizar Valores: {str(e)}")

        return []
